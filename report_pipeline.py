"""
CJ banner report — logic extracted from report_thang3.ipynb.
"""
from __future__ import annotations

import re
import email.header
from copy import copy as copy_style
from html import unescape
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
import pypff
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

matplotlib.use("Agg")
matplotlib.rcParams["font.family"] = ["DejaVu Sans", "Arial Unicode MS", "sans-serif"]

TZ = ZoneInfo("Asia/Ho_Chi_Minh")
TEAM_SENDERS = {"DINH THI HA VI", "NGUYEN HUY", "NGUYEN THIEN HUY"}


def parse_header(headers_str: str, header_name: str) -> str:
    if not headers_str:
        return ""
    pattern = re.compile(
        rf"^{re.escape(header_name)}:\s*(.*?)(?=\n\S|\Z)",
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    )
    m = pattern.search(headers_str)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def decode_mime_header(raw: str) -> str:
    if not raw:
        return ""
    try:
        parts = email.header.decode_header(raw)
        decoded = []
        for part, charset in parts:
            if isinstance(part, bytes):
                decoded.append(part.decode(charset or "utf-8", errors="replace"))
            else:
                decoded.append(part)
        return " ".join(decoded)
    except Exception:
        return raw


def detect_charset(html_bytes: bytes) -> str:
    head = html_bytes[:1024] if len(html_bytes) > 1024 else html_bytes
    m = re.search(rb"charset=([A-Za-z0-9_\-]+)", head, re.IGNORECASE)
    if m:
        charset = m.group(1).decode("ascii", errors="replace").lower()
        if charset in ("ks_c_5601-1987", "ks_c_5601", "ksc5601", "euc-kr", "euckr"):
            return "euc-kr"
        return charset
    return "utf-8"


def decode_body(raw_bytes: bytes | str) -> tuple[str, str]:
    if not raw_bytes:
        return "", "utf-8"
    if isinstance(raw_bytes, str):
        return raw_bytes, "utf-8"
    charset = detect_charset(raw_bytes)
    try:
        return raw_bytes.decode(charset, errors="replace"), charset
    except (UnicodeDecodeError, LookupError):
        return raw_bytes.decode("utf-8", errors="replace"), "utf-8"


def html_to_text(html_str: str) -> str:
    if not html_str:
        return ""
    soup = BeautifulSoup(html_str, "html.parser")
    return soup.get_text(separator=" ", strip=True)


def extract_messages(folder, folder_path: str = "") -> list[dict]:
    messages = []
    name = folder.name or "(unnamed)"
    current_path = f"{folder_path}/{name}" if folder_path else name

    for i in range(folder.number_of_sub_messages):
        msg = folder.get_sub_message(i)
        try:
            subject = msg.subject or ""
            sender_name = msg.sender_name or ""
            conversation_topic = msg.conversation_topic or ""

            delivery_time = None
            try:
                delivery_time = msg.delivery_time
            except Exception:
                pass

            headers = msg.transport_headers or ""
            if isinstance(headers, bytes):
                headers = headers.decode("utf-8", errors="replace")

            from_header = decode_mime_header(parse_header(headers, "From"))
            to_header = decode_mime_header(parse_header(headers, "To"))
            message_id = parse_header(headers, "Message-ID") or parse_header(headers, "Message-Id")
            in_reply_to = parse_header(headers, "In-Reply-To")
            references = parse_header(headers, "References")

            body_html = ""
            body_charset = "utf-8"
            plain_text = ""

            try:
                raw = msg.html_body
                if raw:
                    body_html, body_charset = decode_body(raw)
            except Exception:
                pass

            try:
                pt = msg.plain_text_body
                if pt:
                    plain_text = pt.decode("utf-8", errors="replace") if isinstance(pt, bytes) else pt
            except Exception:
                pass

            body_text = plain_text if plain_text.strip() else html_to_text(body_html)

            num_attachments = 0
            attachment_names = []
            try:
                num_attachments = msg.number_of_attachments
                for a in range(num_attachments):
                    att = msg.get_attachment(a)
                    att_name = ""
                    try:
                        att_name = att.name or ""
                    except Exception:
                        pass
                    attachment_names.append(att_name or f"attachment_{a}")
            except Exception:
                pass

            messages.append(
                {
                    "subject": subject,
                    "conversation_topic": conversation_topic,
                    "sender_name": sender_name,
                    "from": from_header,
                    "to": to_header,
                    "date": delivery_time,
                    "message_id": message_id,
                    "in_reply_to": in_reply_to,
                    "references": references,
                    "body_html": body_html,
                    "body_text": body_text,
                    "body_charset": body_charset,
                    "num_attachments": num_attachments,
                    "attachment_names": "; ".join(attachment_names) if attachment_names else "",
                }
            )
        except Exception as e:
            messages.append(
                {
                    "subject": f"[ERROR: {e}]",
                    "conversation_topic": "",
                    "sender_name": "",
                    "from": "",
                    "to": "",
                    "date": None,
                    "message_id": "",
                    "in_reply_to": "",
                    "references": "",
                    "body_html": "",
                    "body_text": "",
                    "body_charset": "",
                    "num_attachments": 0,
                    "attachment_names": "",
                }
            )

    for i in range(folder.number_of_sub_folders):
        messages.extend(extract_messages(folder.get_sub_folder(i), current_path))
    return messages


def load_dataframe_from_pst(pst_path: str, date_from: str | None, date_to: str | None) -> pd.DataFrame:
    pst = pypff.file()
    pst.open(pst_path)
    root = pst.get_root_folder()
    all_messages = extract_messages(root)
    pst.close()

    df = pd.DataFrame(all_messages)
    df["date"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(TZ)

    if date_from:
        df = df[df["date"] >= date_from]
    if date_to:
        df = df[df["date"] < pd.Timestamp(date_to, tz=TZ) + pd.Timedelta(days=1)]
    return df.sort_values("date").reset_index(drop=True)


def is_team_sender(sender_name: str) -> bool:
    if not sender_name:
        return False
    name = sender_name.upper().replace("님", "").replace(" IN TEAMS", "").strip()
    for ts in TEAM_SENDERS:
        if ts in name:
            return True
    return False


def classify_email(subject, body_text, body_html, sender_name: str = "") -> str:
    subj = subject or ""
    text = body_text or ""
    html = body_html or ""
    sender = sender_name or ""
    team = is_team_sender(sender)

    is_reply = bool(re.search(r"(?:RE|Re|re|FW|Fw|fw)\s*:", subj))

    if is_reply and team:
        sent_patterns = [
            "완료하여 전달드립니다",
            "재전달드립니다",
            "배너를 제작하여",
            "제작 완료",
        ]
        for pat in sent_patterns:
            if pat in text or pat in html:
                return "SENT_REPLY"

    completion_prefixes = ["[완료", "[제작완료", "[제작 완료", "[수정완료", "[수정 완료"]
    for prefix in completion_prefixes:
        if subj.startswith(prefix):
            return "SENT_REPLY"

    if is_reply and "[완료" in subj and team:
        return "SENT_REPLY"

    if "[CJ온스타일]" in subj and ("유형타입" in text or "유형타입" in html):
        return "CLIENT_FEEDBACK" if is_reply else "CAS"

    if "라이브쇼" in subj or "방송일정" in text or "PGM" in text:
        return "CLIENT_FEEDBACK" if is_reply else "LIVE"

    if "제작" in subj or "[요청]" in subj:
        return "CLIENT_FEEDBACK" if is_reply else "MANUAL"

    if is_reply:
        return "CLIENT_FEEDBACK"

    return "OTHER"


def parse_cas_html_table(body_html: str) -> dict:
    soup = BeautifulSoup(body_html, "html.parser")
    form_data = {}
    rows = soup.find_all("tr")
    for row in rows:
        tds = row.find_all("td")
        if len(tds) >= 2:
            label = tds[0].get_text(strip=True)
            value = tds[1].get_text(strip=True)
            if label:
                form_data[label] = unescape(value)
    return form_data


def parse_cas_plain_text(body_text: str) -> dict:
    data = {}
    m = re.search(r"상품코드/명\s+(.+?)(?=유형타입)", body_text, re.DOTALL)
    if m:
        data["상품코드/명"] = m.group(1).strip()
    m = re.search(r"유형타입-H1\s+(.+?)(?=유형타입-550)", body_text, re.DOTALL)
    if m:
        data["유형타입-H1"] = m.group(1).strip()
    m = re.search(r"유형타입-550\s+(.+?)(?=파일링크|요청사항)", body_text, re.DOTALL)
    if m:
        data["유형타입-550"] = m.group(1).strip()
    m = re.search(r"완료요청일\s+(\d{4}-\d{2}-\d{2})", body_text)
    if m:
        data["완료요청일"] = m.group(1)
    return data


def extract_cas_info(row) -> pd.Series:
    body_html = row["body_html"] or ""
    body_text = row["body_text"] or ""

    if "<td" in body_html and "유형타입" in body_html:
        data = parse_cas_html_table(body_html)
    else:
        data = parse_cas_plain_text(body_text)

    code_raw = data.get("상품코드/명", "")
    product_code = ""
    if code_raw:
        parts = code_raw.split("/")
        product_code = parts[0].strip().split()[0] if parts else ""
    if not product_code:
        m = re.search(r"(M\d{7}|\d{10,})", body_text)
        if m:
            product_code = m.group(1)

    h1_raw = data.get("유형타입-H1", "")
    h1_variant = ""
    h1_active = False
    if h1_raw and "제작 안함" not in h1_raw and h1_raw.strip():
        h1_active = True
        m = re.match(r"([A-Z]\d)", h1_raw)
        h1_variant = m.group(1) if m else h1_raw.split("(")[0].strip()

    h550_raw = data.get("유형타입-550", "")
    h550_variant = ""
    h550_active = False
    if h550_raw and "제작 안함" not in h550_raw and h550_raw.strip():
        h550_active = True
        m = re.match(r"([A-Z]\d)", h550_raw)
        h550_variant = m.group(1) if m else h550_raw.split("(")[0].strip()

    request_number = int(h1_active) + int(h550_active)
    if request_number == 0:
        request_number = 1

    if h1_active and h550_active:
        banner_code = f"{product_code} (H1:{h1_variant}, 550:{h550_variant})"
    elif h1_active:
        banner_code = f"{product_code} (H1:{h1_variant})"
    elif h550_active:
        banner_code = f"{product_code} (550:{h550_variant})"
    else:
        banner_code = product_code

    deadline = data.get("완료요청일", "")

    return pd.Series(
        {
            "product_code": product_code,
            "h1_variant": h1_variant,
            "h550_variant": h550_variant,
            "h1_active": h1_active,
            "h550_active": h550_active,
            "request_number": request_number,
            "banner_code": banner_code,
            "deadline": deadline,
        }
    )


def check_urgent(subject, body_text) -> bool:
    subj = subject or ""
    text = body_text or ""

    if "긴급" in subj:
        return True
    if re.search(r"급(?![여식행속수진냉])", subj):
        return True
    if "urgent" in subj.lower():
        return True

    template_exclude = "긴급/당일 요청은 대응이 제한될"
    body_clean = text.replace(template_exclude, "")

    if "긴급" in body_clean:
        return True
    if "urgent" in body_clean.lower():
        return True
    if re.search(r"(?<![발공보납지])급(?![여식행속수진냉])", body_clean):
        return True

    return False


def build_feedback_map(df_all: pd.DataFrame, tasks_df: pd.DataFrame) -> tuple[dict, list[str]]:
    lines: list[str] = []
    sent_replies = df_all[df_all["email_type"] == "SENT_REPLY"]
    task_msgids = set(tasks_df["message_id"].dropna())
    fb_emails = df_all[df_all["email_type"] == "CLIENT_FEEDBACK"]

    sent_to_task = {}
    for _, sr in sent_replies.iterrows():
        reply_to = sr["in_reply_to"]
        if reply_to and reply_to in task_msgids:
            sent_to_task[sr["message_id"]] = reply_to
        else:
            for ref in (sr["references"] or "").split():
                ref = ref.strip()
                if ref in task_msgids:
                    sent_to_task[sr["message_id"]] = ref
                    break

    feedback_counts: dict[str, int] = {}

    for _, fb in fb_emails.iterrows():
        reply_to = fb["in_reply_to"]
        matched_task = None

        if reply_to and reply_to in task_msgids:
            matched_task = reply_to
        elif reply_to and reply_to in sent_to_task:
            matched_task = sent_to_task[reply_to]
        else:
            for ref in (fb["references"] or "").split():
                ref = ref.strip()
                if ref in task_msgids:
                    matched_task = ref
                    break
                if ref in sent_to_task:
                    matched_task = sent_to_task[ref]
                    break

        if matched_task:
            feedback_counts[matched_task] = feedback_counts.get(matched_task, 0) + 1

    method1_count = len(feedback_counts)

    code_to_tasks: dict[str, list] = {}
    for _, task in tasks_df.iterrows():
        pc = task.get("product_code", "")
        if pc:
            code_to_tasks.setdefault(pc, []).append(task["message_id"])

    method2_new = 0
    for _, fb in fb_emails.iterrows():
        subj = fb["subject"] or ""
        m = re.search(r"\[완료[:\s]*\s*(M?\d+)", subj)
        if m:
            code = m.group(1)
            if code in code_to_tasks:
                for tmid in code_to_tasks[code]:
                    if tmid not in feedback_counts:
                        feedback_counts[tmid] = 1
                        method2_new += 1
                    else:
                        feedback_counts[tmid] += 1

    fb_topics = set(fb_emails["conversation_topic"].dropna().unique())
    method3_new = 0
    for _, task in tasks_df[tasks_df["email_type"].isin(["LIVE", "MANUAL"])].iterrows():
        topic = task["conversation_topic"]
        if topic and topic in fb_topics:
            mid = task["message_id"]
            if mid not in feedback_counts:
                count = len(fb_emails[fb_emails["conversation_topic"] == topic])
                feedback_counts[mid] = count
                method3_new += 1

    lines.append(f"  Method 1 (msg chain): {method1_count} tasks")
    lines.append(f"  Method 2 (product code): +{method2_new} new tasks")
    lines.append(f"  Method 3 (topic): +{method3_new} new tasks")

    return feedback_counts, lines


def assign_task_ids(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["date_str"] = df["date"].dt.strftime("%m%d")
    task_ids = []
    daily_counter: dict[str, int] = {}
    for _, row in df.iterrows():
        d = row["date_str"]
        daily_counter[d] = daily_counter.get(d, 0) + 1
        task_ids.append(f"{d}-{daily_counter[d]:02d}")
    df["task_id"] = task_ids
    return df


def build_expanded_df(tasks_df: pd.DataFrame) -> pd.DataFrame:
    expanded_rows = []
    for _, row in tasks_df.iterrows():
        if row["email_type"] == "CAS" and row.get("h1_active") and row.get("h550_active"):
            h1 = row.copy()
            h1["banner_type"] = "H1"
            h1["banner_code"] = f"{row['product_code']} (H1:{row['h1_variant']})"
            h1["request_number"] = 1
            expanded_rows.append(h1)
            h550 = row.copy()
            h550["banner_type"] = "550"
            h550["banner_code"] = f"{row['product_code']} (550:{row['h550_variant']})"
            h550["request_number"] = 1
            expanded_rows.append(h550)
        else:
            single = row.copy()
            if row["email_type"] == "CAS":
                single["banner_type"] = "H1" if row.get("h1_active") else "550"
            else:
                single["banner_type"] = row["email_type"]
            expanded_rows.append(single)

    exp_df = pd.DataFrame(expanded_rows).reset_index(drop=True)
    exp_df["request_number"] = 1

    new_ids = []
    suffix_map: dict[str, Any] = {}
    for _, row in exp_df.iterrows():
        tid = row["task_id"]
        if tid not in suffix_map:
            suffix_map[tid] = iter("abcdefgh")
        count = (exp_df["task_id"] == tid).sum()
        if count > 1:
            new_ids.append(f"{tid}{next(suffix_map[tid])}")
        else:
            new_ids.append(tid)
    exp_df["task_id"] = new_ids
    return exp_df


def run_pipeline(pst_path: str, date_from: str | None, date_to: str | None) -> dict[str, Any]:
    df = load_dataframe_from_pst(pst_path, date_from, date_to)

    df["email_type"] = [
        classify_email(r["subject"], r["body_text"], r["body_html"], r["sender_name"])
        for _, r in df.iterrows()
    ]

    task_types = ["CAS", "LIVE", "MANUAL"]
    tasks_df = df[df["email_type"].isin(task_types)].copy().reset_index(drop=True)

    for col, default in (
        ("product_code", ""),
        ("h1_variant", ""),
        ("h550_variant", ""),
        ("h1_active", False),
        ("h550_active", False),
        ("request_number", 1),
        ("banner_code", ""),
        ("deadline", ""),
    ):
        if col not in tasks_df.columns:
            tasks_df[col] = default

    cas_mask = tasks_df["email_type"] == "CAS"
    if cas_mask.sum() > 0:
        cas_info = tasks_df[cas_mask].apply(extract_cas_info, axis=1)
        for col in cas_info.columns:
            tasks_df.loc[cas_mask, col] = cas_info[col].values

    live_mask = tasks_df["email_type"] == "LIVE"
    tasks_df.loc[live_mask, "request_number"] = 1
    tasks_df.loc[live_mask, "banner_code"] = tasks_df.loc[live_mask, "subject"]
    tasks_df.loc[live_mask, "product_code"] = ""
    tasks_df.loc[live_mask, "deadline"] = ""

    manual_mask = tasks_df["email_type"] == "MANUAL"
    tasks_df.loc[manual_mask, "request_number"] = 1
    tasks_df.loc[manual_mask, "banner_code"] = "Manual"
    tasks_df.loc[manual_mask, "product_code"] = ""
    tasks_df.loc[manual_mask, "deadline"] = ""

    tasks_df["request_number"] = tasks_df["request_number"].fillna(1).astype(int)

    tasks_df["urgent"] = [
        check_urgent(r["subject"], r["body_text"]) for _, r in tasks_df.iterrows()
    ]

    feedback_counts, fb_lines = build_feedback_map(df, tasks_df)
    tasks_df["feedback"] = tasks_df["message_id"].isin(feedback_counts)
    tasks_df["revision_count"] = tasks_df["message_id"].map(feedback_counts).fillna(0).astype(int)

    tasks_df = tasks_df.sort_values("date").reset_index(drop=True)
    tasks_df = assign_task_ids(tasks_df)

    report = tasks_df[
        [
            "task_id",
            "date",
            "email_type",
            "request_number",
            "banner_code",
            "urgent",
            "feedback",
            "revision_count",
            "sender_name",
            "from",
        ]
    ].copy()

    report.insert(0, "STT", range(1, len(report) + 1))
    report = report.rename(
        columns={
            "task_id": "Task",
            "date": "Date",
            "email_type": "Type",
            "request_number": "Request Number",
            "banner_code": "Banner Code",
            "urgent": "Urgent",
            "feedback": "Feedback",
            "revision_count": "Revision Count",
            "sender_name": "Sender",
            "from": "Sender Email",
        }
    )

    report["Date"] = report["Date"].dt.strftime("%d/%m/%Y")
    report["Urgent"] = report["Urgent"].map({True: "Yes", False: ""})
    report["Feedback"] = report["Feedback"].map({True: "Yes", False: ""})
    report["Revision Count"] = report["Revision Count"].apply(lambda x: x if x > 0 else "")
    report["Banner Code"] = report["Banner Code"].fillna("").str[:80]

    cas_tasks = tasks_df[tasks_df["email_type"] == "CAS"]
    stats = {
        "Total Tasks": len(tasks_df),
        "Total Banners": int(tasks_df["request_number"].sum()),
        "CAS Tasks": len(cas_tasks),
        "CAS Banners (H1)": int(cas_tasks["h1_active"].sum()) if "h1_active" in cas_tasks.columns else 0,
        "CAS Banners (550)": int(cas_tasks["h550_active"].sum()) if "h550_active" in cas_tasks.columns else 0,
        "LIVE Tasks/Banners": int((tasks_df["email_type"] == "LIVE").sum()),
        "MANUAL Tasks/Banners": int((tasks_df["email_type"] == "MANUAL").sum()),
        "Urgent Tasks": int(tasks_df["urgent"].sum()),
        "Tasks with Feedback": int(tasks_df["feedback"].sum()),
        "Total Revisions": int(tasks_df["revision_count"].sum()),
    }
    stats_df = pd.DataFrame(list(stats.items()), columns=["Metric", "Value"])

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))

    banner_counts = {
        "CAS H1": stats.get("CAS Banners (H1)", 0),
        "CAS 550": stats.get("CAS Banners (550)", 0),
        "LIVE": stats.get("LIVE Tasks/Banners", 0),
        "MANUAL": stats.get("MANUAL Tasks/Banners", 0),
    }
    colors_bar = ["#4472C4", "#5B9BD5", "#ED7D31", "#70AD47"]
    bars = axes[0].bar(banner_counts.keys(), banner_counts.values(), color=colors_bar)
    axes[0].set_title("Total Banners by Type", fontsize=14, fontweight="bold")
    axes[0].set_ylabel("Count")
    for bar, val in zip(bars, banner_counts.values()):
        axes[0].text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 2,
            str(int(val)),
            ha="center",
            fontweight="bold",
        )

    tasks_df_plot = tasks_df.copy()
    tasks_df_plot["day"] = tasks_df_plot["date"].dt.date
    daily = tasks_df_plot.groupby("day").size()
    axes[1].plot(daily.index, daily.values, marker="o", color="#4472C4", linewidth=2)
    axes[1].set_title("Tasks per Day", fontsize=14, fontweight="bold")
    axes[1].set_ylabel("Task Count")
    axes[1].tick_params(axis="x", rotation=45)

    top_senders = tasks_df["sender_name"].value_counts().head(10)
    axes[2].barh(top_senders.index[::-1], top_senders.values[::-1], color="#4472C4")
    axes[2].set_title("Top 10 Senders", fontsize=14, fontweight="bold")
    axes[2].set_xlabel("Task Count")

    plt.tight_layout()

    exp_df = build_expanded_df(tasks_df)

    prefix = tasks_df["date"].min().strftime("%m%d") if len(tasks_df) else "report"

    return {
        "df": df,
        "tasks_df": tasks_df,
        "report": report,
        "stats_df": stats_df,
        "stats": stats,
        "fig": fig,
        "feedback_debug_lines": fb_lines,
        "exp_df": exp_df,
        "file_prefix": prefix,
    }


def excel_report_to_bytes(report: pd.DataFrame, stats_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    urgent_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    feedback_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = list(report.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(report.iterrows(), 2):
        is_urgent = row["Urgent"] == "Yes"
        is_feedback = row["Feedback"] == "Yes"

        for col_idx, col_name in enumerate(headers, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

            if is_urgent:
                cell.fill = urgent_fill
            elif is_feedback:
                cell.fill = feedback_fill

    col_widths = {
        "STT": 6,
        "Date": 12,
        "Task": 10,
        "Type": 10,
        "Request Number": 10,
        "Banner Code": 45,
        "Urgent": 8,
        "Feedback": 10,
        "Revision Count": 12,
        "Sender": 22,
        "Sender Email": 30,
    }
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws1.cell(row=1, column=col_idx).column_letter
        ws1.column_dimensions[col_letter].width = col_widths.get(header, 15)

    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Statistics")
    for row_idx, (_, row) in enumerate(stats_df.iterrows(), 1):
        ws2.cell(row=row_idx, column=1, value=row["Metric"]).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=row["Value"])
        ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2).border = thin_border
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_banners_to_bytes(exp_df: pd.DataFrame) -> bytes:
    exp_report = exp_df[
        [
            "task_id",
            "date",
            "email_type",
            "banner_type",
            "request_number",
            "banner_code",
            "urgent",
            "feedback",
            "revision_count",
            "sender_name",
            "from",
        ]
    ].copy()
    exp_report.insert(0, "STT", range(1, len(exp_report) + 1))
    exp_report = exp_report.rename(
        columns={
            "task_id": "Task",
            "date": "Date",
            "email_type": "Type",
            "banner_type": "Banner Type",
            "request_number": "Request Number",
            "banner_code": "Banner Code",
            "urgent": "Urgent",
            "feedback": "Feedback",
            "revision_count": "Revision Count",
            "sender_name": "Sender",
            "from": "Sender Email",
        }
    )
    exp_report["Date"] = exp_report["Date"].dt.strftime("%d/%m/%Y")
    exp_report["Urgent"] = exp_report["Urgent"].map({True: "Yes", False: ""})
    exp_report["Feedback"] = exp_report["Feedback"].map({True: "Yes", False: ""})
    exp_report["Revision Count"] = exp_report["Revision Count"].apply(lambda x: x if x > 0 else "")
    exp_report["Banner Code"] = exp_report["Banner Code"].fillna("").str[:80]

    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Banners"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    urgent_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    feedback_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = list(exp_report.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(exp_report.iterrows(), 2):
        is_urgent = row["Urgent"] == "Yes"
        is_feedback = row["Feedback"] == "Yes"
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if is_urgent:
                cell.fill = urgent_fill
            elif is_feedback:
                cell.fill = feedback_fill

    col_widths = {
        "STT": 6,
        "Task": 10,
        "Date": 12,
        "Type": 8,
        "Banner Type": 12,
        "Request Number": 10,
        "Banner Code": 35,
        "Urgent": 8,
        "Feedback": 10,
        "Revision Count": 12,
        "Sender": 22,
        "Sender Email": 30,
    }
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(header, 15)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb2.save(buf)
    return buf.getvalue()


def export_operational_template(
    exp_df: pd.DataFrame,
    template_path: str,
    md_lookup_path: str,
) -> tuple[bytes, str]:
    """Map exp_df to 운영접수리스트 template. Returns (xlsx bytes, summary text)."""
    _lookup = pd.read_excel(md_lookup_path, sheet_name="Sheet2", header=1)
    _lookup["_name"] = _lookup["이름"].apply(
        lambda n: re.sub(r"님\s*$", "", str(n)).strip() if pd.notna(n) else ""
    )
    MD_TEAM_MAP = dict(zip(_lookup["_name"], _lookup["팀"]))

    wb = load_workbook(template_path)
    ws = wb["raw data"]

    template_styles = {}
    for col in range(2, 13):
        src_cell = ws.cell(row=3, column=col)
        template_styles[col] = {
            "font": copy_style(src_cell.font),
            "alignment": copy_style(src_cell.alignment),
            "fill": copy_style(src_cell.fill),
            "border": copy_style(src_cell.border),
            "number_format": src_cell.number_format,
        }

    for row in range(3, 5):
        for col in range(1, 14):
            ws.cell(row=row, column=col).value = None

    def clean_md_name(name):
        if not name:
            return ""
        name = re.sub(r"\[.*?\]", "", name)
        name = re.sub(r"님\s*$", "", name)
        return name.strip()

    TYPE_TO_WORK = {"CAS": "시스템배너", "LIVE": "라이브배너", "MANUAL": "배너"}

    num_rows = len(exp_df)
    unmatched_names = set()

    for idx, (_, row) in enumerate(exp_df.iterrows()):
        r = idx + 3

        md_name = clean_md_name(row["sender_name"])
        team = MD_TEAM_MAP.get(md_name, "")
        if not team and md_name:
            unmatched_names.add(md_name)

        ws.cell(r, 2, value=idx + 1)
        ws.cell(r, 3, value=team)
        ws.cell(r, 4, value=md_name)
        ws.cell(r, 5, value="긴급" if row.get("urgent") else "일반")
        ws.cell(r, 6, value="수정" if row.get("feedback") else "신규")
        ws.cell(r, 7, value=TYPE_TO_WORK.get(row["email_type"], "배너"))
        ws.cell(r, 8, value=row.get("subject", ""))
        note = row.get("banner_code", "") if row["email_type"] == "CAS" else ""
        ws.cell(r, 9, value=note)
        ws.cell(r, 10, value=row.get("deadline", "") or "")
        ws.cell(r, 11, value="완료")
        ws.cell(r, 12, value="")

        for col in range(2, 13):
            cell = ws.cell(r, col)
            st = template_styles[col]
            cell.font = copy_style(st["font"])
            cell.alignment = copy_style(st["alignment"])
            cell.fill = copy_style(st["fill"])
            cell.border = copy_style(st["border"])
            cell.number_format = st["number_format"]

    last_row = num_rows + 2
    ws.auto_filter.ref = f"B2:L{last_row}"

    END = last_row + 20
    ws_s = wb["summary"]
    for row in range(4, 32):
        for c in ["F", "G"]:
            cell = ws_s[f"{c}{row}"]
            if cell.value and isinstance(cell.value, str) and "COUNTIFS" in cell.value:
                cell.value = (
                    cell.value.replace("$C$1:$C$106", f"$C$1:$C${END}")
                    .replace("$F$1:$F$106", f"$F$1:$F${END}")
                )

    ws_d = wb["세부접수내역"]
    for row in range(4, 32):
        for col in range(3, 19):
            cell = ws_d.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "COUNTIFS" in cell.value:
                cell.value = (
                    cell.value.replace("$C$3:$C$252", f"$C$3:$C${END}")
                    .replace("$G$3:$G$252", f"$G$3:$G${END}")
                )

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    lines = [
        f"Wrote {num_rows} rows",
        f"COUNTIFS end row ~ {END}",
    ]
    n_new = (exp_df["feedback"] != True).sum()
    n_mod = (exp_df["feedback"] == True).sum()
    lines.append(f"신규={n_new}, 수정={n_mod}")
    if unmatched_names:
        lines.append(f"⚠ {len(unmatched_names)} senders not in MD lookup:")
        for n in sorted(unmatched_names):
            cnt = (exp_df["sender_name"].apply(clean_md_name) == n).sum()
            lines.append(f"  {n} ({cnt} rows)")

    return data, "\n".join(lines)
